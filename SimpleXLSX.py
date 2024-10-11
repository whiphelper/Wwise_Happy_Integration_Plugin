import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from colorsys import rgb_to_hls, hls_to_rgb
from openpyxl.styles.colors import COLOR_INDEX
from xml.etree.ElementTree import QName, fromstring
from globals import *


class ThemeColorConverter:  # 这部分代码转载于：https://blog.csdn.net/as604049322/article/details/134470419  本内容依据“CC BY-SA 4.0”许可证进行授权。要查看该许可证，可访问https://creativecommons.org/licenses/by-sa/4.0/
    RGBMAX = 0xff
    HLSMAX = 240

    def __init__(self, wb):
        self.colors = self.get_theme_colors(wb)

    @staticmethod
    def tint_luminance(tint, lum):
        if tint < 0:
            return int(round(lum * (1.0 + tint)))
        return int(round((ThemeColorConverter.HLSMAX - lum) * tint)) + lum

    @staticmethod
    def ms_hls_to_rgb(hue, lightness=None, saturation=None):
        if lightness is None:
            hue, lightness, saturation = hue
        hlsmax = ThemeColorConverter.HLSMAX
        return hls_to_rgb(hue / hlsmax, lightness / hlsmax, saturation / hlsmax)

    @staticmethod
    def rgb_to_hex(red, green=None, blue=None):
        if green is None:
            red, green, blue = red
        return '{:02X}{:02X}{:02X}'.format(
            int(red * ThemeColorConverter.RGBMAX),
            int(green * ThemeColorConverter.RGBMAX),
            int(blue * ThemeColorConverter.RGBMAX)
        )

    @staticmethod
    def rgb_to_ms_hls(red, green=None, blue=None):
        if green is None:
            if isinstance(red, str):
                if len(red) > 6:
                    red = red[-6:]  # Ignore preceding '#' and alpha values
                rgbmax = ThemeColorConverter.RGBMAX
                blue = int(red[4:], 16) / rgbmax
                green = int(red[2:4], 16) / rgbmax
                red = int(red[0:2], 16) / rgbmax
            else:
                red, green, blue = red
        h, l, s = rgb_to_hls(red, green, blue)
        hlsmax = ThemeColorConverter.HLSMAX
        return (int(round(h * hlsmax)), int(round(l * hlsmax)),
                int(round(s * hlsmax)))

    @staticmethod
    def get_theme_colors(wb):
        from xml.etree.ElementTree import QName, fromstring
        xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        root = fromstring(wb.loaded_theme)
        themeEl = root.find(QName(xlmns, 'themeElements').text)
        colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
        firstColorScheme = colorSchemes[0]
        colors = []
        for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
            accent = firstColorScheme.find(QName(xlmns, c).text)
            for i in list(accent):
                if 'window' in i.attrib['val']:
                    colors.append(i.attrib['lastClr'])
                else:
                    colors.append(i.attrib['val'])
        return colors

    def theme_and_tint_to_rgb(self, theme, tint):
        rgb = self.colors[theme]
        h, l, s = self.rgb_to_ms_hls(rgb)
        return self.rgb_to_hex(self.ms_hls_to_rgb(h, self.tint_luminance(tint, l), s))


class SimpleXLSX:
    def __init__(self):
        self.wb = None
        self.ws = None

    def ReadXLSX(self, xlsxPath, sheetIndex: int):
        self.wb = openpyxl.load_workbook(xlsxPath)
        self.ws = self.wb.worksheets[sheetIndex]

    def CreateOneSheetXLSX(self, xlsxPath, sheetName):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheetName
        self.wb.save(xlsxPath)

    def CreateMultiSheetXLSX(self, xlsxPath, sheetNameList: list):
        if len(sheetNameList) != 0:
            self.wb = Workbook()
            for i in range(len(sheetNameList)):
                self.wb.create_sheet(sheetNameList[i], i)

            # 移除默认的空sheet
            EmptySheet = self.wb.worksheets[len(sheetNameList)]
            self.wb.remove(EmptySheet)

            self.wb.save(xlsxPath)

    def WriteCell(self, Cell, Str):
        self.ws[Cell] = Str

    def WriteCell_TextAndColor(self, Cell, Str, textColor, bgColor):
        self.ws[Cell] = Str
        if len(textColor) == 7:
            textColor = self.ColorConvert_Hex_to_ARGB(textColor)
        if len(bgColor) == 7:
            bgColor = self.ColorConvert_Hex_to_ARGB(bgColor)
        self.ws[Cell].font = Font(color=textColor)
        self.ws[Cell].fill = PatternFill("solid", fgColor=bgColor)

    def ColorConvert_Hex_to_ARGB(self, hexValue):
        if type(hexValue) is not str:
            return "ffffffff"
        else:
            if len(hexValue) != 7:
                return "ffffffff"
            else:
                head = hexValue[0:1]
                tail = hexValue[1:]
                if head != "#":
                    return "ffffffff"
                else:
                    ARGBValue = "ff" + tail
                    return ARGBValue

    def GetSheetNameList(self, XLSXPath):
        workbook = openpyxl.load_workbook(XLSXPath)
        sheetNames = workbook.sheetnames
        for eachName in sheetNames:
            LOG.debug(eachName)

    def GetCellValue(self, Cell):
        return str(self.ws[Cell].value)

    def AdjustColumnWidth(self):  # 方法转载于：https://blog.csdn.net/crammy/article/details/120469646  本内容依据“CC BY-SA 4.0”许可证进行授权。要查看该许可证，可访问https://creativecommons.org/licenses/by-sa/4.0/
        lks = []
        for i in range(1, self.ws.max_column + 1):
            lk = 1
            for j in range(1, self.ws.max_row + 1):
                sz = self.ws.cell(row=j, column=i).value
                if isinstance(sz, str):
                    lk1 = len(sz.encode('gbk'))
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1
            lks.append(lk)

        for i in range(1, self.ws.max_column + 1):
            k = get_column_letter(i)
            self.ws.column_dimensions[k].width = lks[i - 1] + 2

    def GetMaxRow(self):  # 方法转载于：https://blog.csdn.net/weixin_44981444/article/details/110502864  本内容依据“CC BY-SA 4.0”许可证进行授权。要查看该许可证，可访问https://creativecommons.org/licenses/by-sa/4.0/
        MaxRowNum = self.ws.max_row
        real_max_row = 0
        while MaxRowNum > 0:
            row_dict = {MaxRowNum.value for MaxRowNum in self.ws[MaxRowNum]}
            if row_dict == {None}:
                MaxRowNum = MaxRowNum - 1
            else:
                real_max_row = MaxRowNum
                break

        return real_max_row

    def FillBGColor(self, Cell, Color):
        self.ws[Cell].fill = PatternFill("solid", fgColor=Color)

    def BoldText(self, Cell):
        self.ws[Cell].font = Font(bold=True)

    def TextColor(self, Cell, Color):
        self.ws[Cell].font = Font(color=Color)

    def TextSize(self, Cell, Size: int):
        self.ws[Cell].font = Font(size=Size)

    def SaveXLSX(self, xlsxPath):
        # self.AdjustColumnWidth()
        self.wb.save(xlsxPath)

    def get_cell_color_Fill(self, cell):
        cell = self.ws[cell]
        color = cell.fill.start_color
        if color.type == "rgb":
            if color.rgb == "00000000":
                return "#FFFFFF"
            else:
                return "#" + color.rgb[2:]
        elif color.type == "indexed":
            color_index = color.indexed
            if color_index is None or color_index < len(COLOR_INDEX):
                raise Exception("Invalid indexed color")
            return COLOR_INDEX[color_index]
        elif color.type == "theme":
            theme_color = ThemeColorConverter(self.wb).theme_and_tint_to_rgb(color.theme, color.tint)
            return "#" + theme_color
        else:
            raise Exception(f"Other type: {color.type}")

    def get_cell_color_Font(self, cell):
        cell = self.ws[cell]
        color = cell.font.color
        if color is not None:
            if color.type == "rgb":
                if color.rgb == "00000000":
                    return "#FFFFFF"
                else:
                    return "#" + color.rgb[2:]
            elif color.type == "indexed":
                color_index = color.indexed
                if color_index is None or color_index < len(COLOR_INDEX):
                    raise Exception("Invalid indexed color")
                return COLOR_INDEX[color_index]
            elif color.type == "theme":
                theme_color = ThemeColorConverter(self.wb).theme_and_tint_to_rgb(color.theme, color.tint)
                return "#" + theme_color
            else:
                raise Exception(f"Other type: {color.type}")
        else:
            return "#000000"
